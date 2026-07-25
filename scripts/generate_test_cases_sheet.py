import os, re, csv

workspace_dir = r"c:\Users\hp\OneDrive\Pictures\Desktop\Web Application"

test_cases = []

# Parse validation tests
validation_file = os.path.join(workspace_dir, "tests", "validation", "test_validation.py")
if os.path.exists(validation_file):
    with open(validation_file, "r", encoding="utf-8") as f:
        current_class = "General"
        for line in f:
            class_match = re.match(r"class\s+(\w+):", line)
            if class_match:
                current_class = class_match.group(1)
            test_match = re.search(r"def\s+(test_([A-Z0-9]+)_([a-zA-Z0-9_]+))\(", line)
            if test_match:
                func_name = test_match.group(1)
                test_id = test_match.group(2)
                desc = test_match.group(3).replace("_", " ")
                test_cases.append({
                    "ID": test_id,
                    "Suite": "Validation (Pytest)",
                    "Component": current_class,
                    "Function Name": func_name,
                    "Description": f"Validate that {desc}"
                })

# Parse E2E Selenium tests
e2e_file = os.path.join(workspace_dir, "tests", "e2e", "test_e2e_full.py")
if os.path.exists(e2e_file):
    with open(e2e_file, "r", encoding="utf-8") as f:
        current_class = "General"
        for line in f:
            class_match = re.match(r"class\s+(\w+):", line)
            if class_match:
                current_class = class_match.group(1)
            test_match = re.search(r"def\s+(test_([A-Z0-9]+)_([a-zA-Z0-9_]+))\(", line)
            if test_match:
                func_name = test_match.group(1)
                test_id = test_match.group(2)
                desc = test_match.group(3).replace("_", " ")
                test_cases.append({
                    "ID": test_id,
                    "Suite": "End-to-End (Selenium)",
                    "Component": current_class,
                    "Function Name": func_name,
                    "Description": f"Verify E2E flow for {desc}"
                })

# Parse Backend Unit tests
unit_file = os.path.join(workspace_dir, "backend", "tests", "unit", "api.test.js")
if os.path.exists(unit_file):
    with open(unit_file, "r", encoding="utf-8") as f:
        current_suite = "Backend Unit"
        for line in f:
            suite_match = re.search(r"describe\('([^']+)'", line)
            if suite_match:
                current_suite = suite_match.group(1)
            test_match = re.search(r"test\('(([A-Z0-9\-]+):\s*([^']+))'", line)
            if test_match:
                full_name = test_match.group(1)
                test_id = test_match.group(2)
                desc = test_match.group(3)
                test_cases.append({
                    "ID": test_id,
                    "Suite": "Backend Unit (Jest)",
                    "Component": current_suite,
                    "Function Name": full_name,
                    "Description": desc
                })

# Parse Load tests
load_file = os.path.join(workspace_dir, "tests", "load", "locustfile.py")
if os.path.exists(load_file):
    with open(load_file, "r", encoding="utf-8") as f:
        current_class = "General"
        last_comment_id = ""
        for line in f:
            class_match = re.match(r"class\s+(\w+)\(", line)
            if class_match:
                current_class = class_match.group(1)
            comment_match = re.search(r"#\s*─\s*(TC-LOAD-\d+)\s*─", line)
            if comment_match:
                last_comment_id = comment_match.group(1)
            test_match = re.search(r"def\s+([a-zA-Z0-9_]+)\(self\):", line)
            if test_match:
                func_name = test_match.group(1)
                desc = test_match.group(3) if len(test_match.groups()) > 2 else test_match.group(1)
                desc = desc.replace("_", " ")
                test_cases.append({
                    "ID": last_comment_id or "LOAD-TASK",
                    "Suite": "Load (Locust)",
                    "Component": current_class,
                    "Function Name": func_name,
                    "Description": f"Simulate user activity: {desc}"
                })

# Write to CSV
csv_file = os.path.join(workspace_dir, "MicrobeVision_AI_Test_Cases.csv")
with open(csv_file, "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=["ID", "Suite", "Component", "Function Name", "Description"])
    writer.writeheader()
    writer.writerows(test_cases)

print(f"Generated CSV sheet with {len(test_cases)} test cases.")

# Try to write to Excel if openpyxl is available
try:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    
    # Headers
    headers = ["ID", "Suite", "Component", "Function Name", "Description"]
    ws.append(headers)
    
    # Styles
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    # Write data
    for tc in test_cases:
        ws.append([tc["ID"], tc["Suite"], tc["Component"], tc["Function Name"], tc["Description"]])
        
    # Apply styling
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            cell.font = Font(name="Segoe UI", size=10)
            if cell.row == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
            else:
                if cell.column in [1, 2]:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    excel_file = os.path.join(workspace_dir, "MicrobeVision_AI_Test_Cases.xlsx")
    wb.save(excel_file)
    print(f"Generated XLSX sheet with {len(test_cases)} test cases.")
except ImportError:
    print("openpyxl not installed, XLSX generation skipped. Only CSV generated.")
